"""Excel export utilities for sash window projects."""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def _ensure_output_dir(path: str | Path) -> Path:
    directory = Path(path)
    directory.mkdir(parents=True, exist_ok=True)
    return directory


def _write_section(ws, start_row: int, title: str, data: Dict[str, object]) -> int:
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    row = start_row + 1
    for key, value in data.items():
        ws.cell(row=row, column=1, value=key.replace("_", " ").title()).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    return row + 1


def _write_materials_table(ws, start_row: int, materials: List[Dict[str, object]]) -> int:
    headers = ["Material Type", "Section", "Length (mm)", "Qty", "Wood Type"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row = start_row + 1
    for item in materials:
        ws.cell(row=row, column=1, value=item["material_type"])
        ws.cell(row=row, column=2, value=item["section"])
        ws.cell(row=row, column=3, value=item["length"])
        ws.cell(row=row, column=4, value=item["qty"])
        ws.cell(row=row, column=5, value=item["wood_type"])
        row += 1

    return row + 2


def generate_excel(project_data: Dict[str, object], export_dir: str = "output") -> str:
    """Create an Excel workbook summarising the project and window details."""

    project = project_data["project"]
    windows = project_data["windows"]
    summary = project_data["summary"]

    output_directory = _ensure_output_dir(export_dir)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    summary_sheet["A1"] = "Skylon Elements – Sash Window Designer"
    summary_sheet["A1"].font = Font(size=14, bold=True)
    summary_sheet["A3"] = "Project"
    summary_sheet["A3"].font = Font(bold=True)
    summary_sheet["B3"] = project["name"]
    summary_sheet["A4"] = "Client"
    summary_sheet["A4"].font = Font(bold=True)
    summary_sheet["B4"] = project["client_name"]
    summary_sheet["A5"] = "Created"
    summary_sheet["A5"].font = Font(bold=True)
    summary_sheet["B5"] = project["created_at"].strftime("%Y-%m-%d %H:%M")

    summary_sheet["A7"] = "Total Timber (mm)"
    summary_sheet["A7"].font = Font(bold=True)
    summary_sheet["B7"] = summary["total_timber"]
    summary_sheet["A8"] = "Total Glass Area (m²)"
    summary_sheet["A8"].font = Font(bold=True)
    summary_sheet["B8"] = summary["total_glass_area"]
    summary_sheet["A9"] = "Window Count"
    summary_sheet["A9"].font = Font(bold=True)
    summary_sheet["B9"] = summary["window_count"]

    for idx, window in enumerate(windows, start=1):
        sheet = workbook.create_sheet(title=f"W-{idx}")
        sheet["A1"] = f"Window {window['window']['name']}"
        sheet["A1"].font = Font(size=13, bold=True)

        row = _write_section(sheet, 3, "Frame", window["window"]["frame"])
        row = _write_section(sheet, row, "Top Sash", window["window"]["sash_top"])
        row = _write_section(sheet, row, "Bottom Sash", window["window"]["sash_bottom"])
        row = _write_section(sheet, row, "Glass", window["glass"])
        row = _write_section(sheet, row, "Bars", window["bars"])

        sheet.cell(row=row, column=1, value="Hardware").font = Font(bold=True, size=12)
        for offset, (key, value) in enumerate(window["hardware"].items(), start=1):
            sheet.cell(row=row + offset, column=1, value=key.replace("_", " ").title()).font = Font(bold=True)
            sheet.cell(row=row + offset, column=2, value=value)
        row += len(window["hardware"]) + 2

        row = _write_materials_table(sheet, row, window["materials"])
        sheet.cell(row=row, column=1, value="Totals").font = Font(bold=True)
        sheet.cell(row=row, column=2, value="Timber Length (mm)").font = Font(bold=True)
        sheet.cell(row=row, column=3, value=window["totals"]["timber_length"])
        sheet.cell(row=row + 1, column=2, value="Glass Area (m²)").font = Font(bold=True)
        sheet.cell(row=row + 1, column=3, value=window["totals"]["glass_area"])

        for column in range(1, 6):
            sheet.column_dimensions[chr(64 + column)].width = 22

    file_path = output_directory / f"{project['name'].replace(' ', '_').lower()}_report.xlsx"
    workbook.save(file_path)
    return str(file_path)
