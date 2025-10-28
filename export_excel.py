"""
Excel export functionality using openpyxl
Creates detailed worksheets for each window plus summary
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import Project, Window
from calculations import get_cutting_list, get_shopping_list


def create_excel_report(project: Project, output_path: str = None) -> str:
    """
    Create Excel workbook with window specifications

    Args:
        project: Project object containing windows
        output_path: Optional output path (default: output/{project_name}.xlsx)

    Returns:
        Path to created Excel file
    """
    if output_path is None:
        output_path = f"output/{project.name.replace(' ', '_')}.xlsx"

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create sheet for each window
    for window in project.windows:
        create_window_sheet(wb, window)

    # Create summary sheet
    create_summary_sheet(wb, project)

    # Save workbook
    wb.save(output_path)
    print(f"Excel report created: {output_path}")

    return output_path


def create_window_sheet(wb: Workbook, window: Window):
    """Create detailed sheet for a single window"""
    ws = wb.create_sheet(title=window.name)

    # Title
    ws['A1'] = f"Window Specification: {window.name}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')

    # Frame Dimensions
    row = 3
    ws[f'A{row}'] = "FRAME DIMENSIONS"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    row += 1
    frame_data = [
        ("Frame Width", f"{window.frame.width:.1f} mm"),
        ("Frame Height", f"{window.frame.height:.1f} mm"),
        ("Jambs Length", f"{window.frame.jambs_length:.1f} mm"),
        ("Head Length", f"{window.frame.head_length:.1f} mm"),
        ("Cill Length", f"{window.frame.cill_length:.1f} mm"),
        ("Ext Head Liner", f"{window.frame.ext_head_liner:.1f} mm"),
        ("Int Head Liner", f"{window.frame.int_head_liner:.1f} mm"),
    ]

    for label, value in frame_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    # Top Sash
    row += 1
    ws[f'A{row}'] = "TOP SASH DIMENSIONS"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    row += 1
    sash_data = [
        ("Sash Width", f"{window.sash_top.width:.1f} mm"),
        ("Sash Height", f"{window.sash_top.height:.1f} mm"),
        ("Height with Horn", f"{window.sash_top.height_with_horn:.1f} mm"),
        ("Stiles Length", f"{window.sash_top.stiles_length:.1f} mm"),
        ("Top Rail Length", f"{window.sash_top.top_rail_length:.1f} mm"),
        ("Meeting Rail Length", f"{window.sash_top.meet_rail_length:.1f} mm"),
    ]

    for label, value in sash_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    # Bottom Sash
    row += 1
    ws[f'A{row}'] = "BOTTOM SASH DIMENSIONS"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    row += 1
    bottom_sash_data = [
        ("Sash Width", f"{window.sash_bottom.width:.1f} mm"),
        ("Sash Height", f"{window.sash_bottom.height:.1f} mm"),
        ("Height with Horn", f"{window.sash_bottom.height_with_horn:.1f} mm"),
        ("Stiles Length", f"{window.sash_bottom.stiles_length:.1f} mm"),
        ("Bottom Rail Length", f"{window.sash_bottom.bottom_rail_length:.1f} mm"),
        ("Meeting Rail Length", f"{window.sash_bottom.meet_rail_length:.1f} mm"),
    ]

    for label, value in bottom_sash_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    # Glass Specifications
    row += 1
    ws[f'A{row}'] = "GLASS SPECIFICATIONS"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    row += 1
    glass_data = [
        ("Top Glass Width", f"{window.glass_top.width:.1f} mm"),
        ("Top Glass Height", f"{window.glass_top.height:.1f} mm"),
        ("Bottom Glass Width", f"{window.glass_bottom.width:.1f} mm"),
        ("Bottom Glass Height", f"{window.glass_bottom.height:.1f} mm"),
        ("Glass Type", window.glass_top.type),
        ("Spacer Color", window.glass_top.spacer_color),
        ("Toughened", "Yes" if window.glass_top.toughened else "No"),
        ("Frosted", "Yes" if window.glass_top.frosted else "No"),
    ]

    for label, value in glass_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    # Hardware & Finish
    row += 1
    ws[f'A{row}'] = "HARDWARE & FINISH"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    row += 1
    hardware_data = [
        ("Paint Color", window.paint_color),
        ("Hardware Finish", window.hardware_finish),
        ("Trickle Vent", window.trickle_vent),
        ("Sash Catches", window.sash_catches),
        ("Cill Extension", f"{window.cill_extension} mm"),
        ("Bars Layout", window.bars_top.layout_type),
    ]

    for label, value in hardware_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    # Cutting List
    row += 2
    ws[f'A{row}'] = "CUTTING LIST"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    row += 1
    headers = ["Section", "Qty", "Length (mm)", "Wood Type"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    cutting_list = get_cutting_list(window)
    row += 1
    for item in cutting_list:
        ws[f'A{row}'] = item['section']
        ws[f'B{row}'] = item['qty']
        ws[f'C{row}'] = item['length']
        ws[f'D{row}'] = item['wood_type']
        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_summary_sheet(wb: Workbook, project: Project):
    """Create summary sheet with totals"""
    ws = wb.create_sheet(title="Summary", index=0)

    # Title
    ws['A1'] = f"Project Summary: {project.name}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')

    ws['A2'] = f"Client: {project.client_name}"
    ws['A3'] = f"Date: {project.created_at.strftime('%Y-%m-%d %H:%M')}"

    # Windows Overview
    row = 5
    ws[f'A{row}'] = "WINDOWS OVERVIEW"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    row += 1
    headers = ["Window", "Frame Size (W x H)", "Paint", "Hardware"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    row += 1
    for window in project.windows:
        ws[f'A{row}'] = window.name
        ws[f'B{row}'] = f"{window.frame.width:.0f} x {window.frame.height:.0f} mm"
        ws[f'C{row}'] = window.paint_color
        ws[f'D{row}'] = window.hardware_finish
        row += 1

    # Total Timber
    row += 2
    ws[f'A{row}'] = "TOTAL TIMBER REQUIREMENTS"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    row += 1
    total_length = 0
    for window in project.windows:
        cutting_list = get_cutting_list(window)
        for item in cutting_list:
            total_length += item['length'] * item['qty']

    ws[f'A{row}'] = "Total Linear Meters"
    ws[f'B{row}'] = f"{total_length / 1000:.2f} m"
    ws[f'B{row}'].font = Font(bold=True)

    # Total Glass
    row += 2
    ws[f'A{row}'] = "TOTAL GLASS"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    row += 1
    total_glass_area = 0
    for window in project.windows:
        total_glass_area += (window.glass_top.width * window.glass_top.height) / 1000000
        total_glass_area += (window.glass_bottom.width * window.glass_bottom.height) / 1000000

    ws[f'A{row}'] = "Total Glass Area"
    ws[f'B{row}'] = f"{total_glass_area:.2f} m²"
    ws[f'B{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = "Total Panes"
    ws[f'B{row}'] = len(project.windows) * 2  # Top and bottom per window
    ws[f'B{row}'].font = Font(bold=True)

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
